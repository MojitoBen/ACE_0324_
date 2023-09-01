import pymysql
from tkinter import Toplevel, Label, Entry, Button, ttk, filedialog, messagebox, PhotoImage
import tkinter as tk
from datetime import datetime, timedelta, date
from PIL import ImageTk, Image
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
import logging
import cv2
import base64

ng_Checked = False
notification = False  
original_result_count = 0
enlarge_view_window = None  # 定義全局變數來存儲現場畫面視窗

logname = 'UI_log/'
logname = logname+"{:%Y-%m-%d}".format(datetime.now())+'.log'
FORMAT = '%(asctime)s %(levelname)s: %(message)s'
logging.basicConfig(level=logging.INFO, filename=logname, filemode='a', format=FORMAT)


def connect_to_database():
    host = "192.168.XX.XXX"
    user = "xxxxxxx"
    password = "xxxxxxx"
    db = "asc_ai"
    charset = "utf8"

    connection = pymysql.connect(host=host, user=user, password=password, db=db, charset=charset)
    return connection

def is_alphanumeric(input_string):
    return input_string.isalnum()

def select_date(selected_date_var, root, query_entry,total_pages_label,tree, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2):
    # 建立小型新視窗
    date_window = Toplevel(root)

    # 設定新視窗的標題
    date_window.title("選擇日期")

    # 設定新視窗的大小
    window_width = 450
    window_height = 450

    # 獲取螢幕的寬度和高度
    screen_width = date_window.winfo_screenwidth()
    screen_height = date_window.winfo_screenheight()

    # 計算窗口的左上角位置
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    # 設定新視窗的位置
    date_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
    date_window.resizable(False, False)
    #date_window.attributes('-topmost', True)  # 保持新視窗在最上層
    
    # 設定選擇日期的起始值
    selected_date_str = selected_date_var.get()
    selected_date_parts = selected_date_str.split(" ~ ")
    if len(selected_date_parts) >= 2:
        start_date_str = selected_date_parts[0]
        end_date_str = selected_date_parts[1]
    else:
        print("selected_date_parts",selected_date_parts)
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=29)
    start_date = datetime.strptime(start_date_str, "%Y/%m/%d").date()
    end_date = datetime.strptime(end_date_str, "%Y/%m/%d").date()
    check_date = end_date + timedelta(days=1)


    # 設定日期選擇框的標籤
    label = Label(date_window, text="選擇日期範圍", font=("Arial", 20))
    label.pack(pady=10)

    # 設定起始日期選擇框
    start_date_label = Label(date_window, text="起始日期", font=("Arial", 20))
    start_date_label.pack()

    start_date_entry = Entry(date_window, font=("Arial", 20))
    start_date_entry.insert(0, start_date.strftime("%Y/%m/%d"))
    start_date_entry.pack()

    # 設定結束日期選擇框
    end_date_label = Label(date_window, text="結束日期", font=("Arial", 20))
    end_date_label.pack()

    end_date_entry = Entry(date_window, font=("Arial", 20))
    end_date_entry.insert(0, check_date.strftime("%Y/%m/%d"))
    end_date_entry.pack()

    def confirm_date():
        try:
            # 解析起始日期和結束日期
            start_date_str = start_date_entry.get()
            end_date_str = end_date_entry.get()

            start_date = datetime.strptime(start_date_str, "%Y/%m/%d").date()
            end_date = datetime.strptime(end_date_str, "%Y/%m/%d").date()
            check_date = end_date + timedelta(days=1)

            # 檢查日期範圍的有效性
            if start_date > end_date:
                messagebox.showerror("錯誤", "起始日期不能大於結束日期")
                logging.warning("起始日期不能大於結束日期")
                return

            # 將日期格式轉換成正確的格式
            selected_date_var.set(f"{start_date.strftime('%Y/%m/%d')} ~ {check_date.strftime('%Y/%m/%d')}")

            # 關閉日期選擇視窗
            date_window.destroy()

            renew(selected_date_var,query_entry,total_pages_label,tree, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2)

        except ValueError:
            messagebox.showerror("錯誤", "請確認日期格式為YYYY/MM/DD")
            logging.warning("日期格式錯誤")


    # 確認按鈕
    confirm_button = Button(date_window, text="確定", font=("Arial", 20), command=confirm_date)
    confirm_button.pack(pady=10)

    # 取消按鈕
    cancel_button = Button(date_window, text="取消", font=("Arial", 20), command=date_window.destroy)
    cancel_button.pack()

def renew(selected_date_var,query_entry,total_pages_label,tree, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2):
    
    # 清除 query_entry 內的資料
    query_entry.delete(0, tk.END)

    # 清空圖像框的內容
    '''
    for widget in img_frame1.winfo_children():
        widget.destroy()
    for widget in img_frame2.winfo_children():
        widget.destroy()
    '''
    for widget in img_frame3.winfo_children():
        widget.destroy()
    for widget in img_frame4.winfo_children():
        widget.destroy()
    
    # 清空標籤的內容
    #var_label2.configure(text="")
    var_label3.configure(text="")
    var_label4.configure(text="")

    # 重新讀取所選日期內的資料庫內容
    results = get_all_results(selected_date_var)
    # 清空表格內容
    tree.delete(*tree.get_children())

    # 更新頁數標籤
    total_pages = (len(results) - 1) // 10 + 1
    current_page = 1
    total_pages_label.config(text=f"{current_page}/{total_pages}")

    # 插入新的資料到表格
    for i, row in enumerate(results):
        # 調整欄位順序和顯示內容
        sn = row[0]
        snap_time = row[1]
        l1_data = row[2]
        cam1_num = row[3]
        result = row[4]

        # 插入資料到表格
        if result == "OK":
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("OK",))
        elif result == "CHECK":
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("CHECK",))
        else:
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("NG",))
        
        if i >= 9:
            break

def switch_page(target_page, total_pages_label, tree, selected_date_var):
    current_page = int(total_pages_label.cget("text").split("/")[0])
    total_pages = int(total_pages_label.cget("text").split("/")[1])
    
    if current_page <= total_pages:
        results = get_all_results(selected_date_var)
    else:
        messagebox.showerror("錯誤", "錯誤")
        logging.error("switch_page錯誤")
        return

    total_pages_label.config(text=f"{target_page}/{total_pages}")
    tree.delete(*tree.get_children())

    for row in results[(target_page - 1) * 10:target_page * 10]:
        sn = row[0]
        snap_time = row[1]
        l1_data = row[2] if len(row) > 2 else ""
        cam1_num = row[3] if len(row) > 3 else ""
        result = row[4] if len(row) > 4 else ""

        if result == "OK":
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("OK",))
        elif result == "CHECK":
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("CHECK",))
        else:
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("NG",))

def query(query_entry, selected_date_var, tree, total_pages_label):
    query_text = query_entry.get()
    if not is_alphanumeric(query_text):
        messagebox.showerror("錯誤", "請輸入英文字母和數字")
        logging.warning("查詢欄位輸入錯誤")
        return

    results = get_all_results(selected_date_var)
    filtered_result = [row for row in results if row[2] == query_text]

    if len(filtered_result) > 0:
        page_size = 10  # 每頁顯示的資料筆數
        current_page = int(total_pages_label.cget("text").split("/")[0])
        total_pages = int(total_pages_label.cget("text").split("/")[1])
        row_num = -1
        index = results.index(filtered_result[0])

        # 計算目標資料所在的頁數與行數
        target_page = (index // page_size) + 1
        target_row = (index % page_size) + 1

        # 切換到目標頁面
        if target_page != current_page:
            switch_page(target_page,total_pages_label, tree, selected_date_var)  # 呼叫切換頁面的函式
        
        tree.see(tree.get_children()[target_row - 1])
        tree.selection_set(tree.get_children()[target_row - 1])
        tree.focus(tree.get_children()[target_row - 1])

        messagebox.showinfo("查詢結果", f"已找到鋼捲：{query_text}\n頁數：{target_page}\n行數：{target_row}")
        logging.info(f"已找到鋼捲：{query_text}頁數：{target_page}行數：{target_row}")
    else:
        messagebox.showerror("錯誤", f"查無此鋼捲：{query_text}")
        logging.warning(f"查無此鋼捲：{query_text}")

def validate(P):
    if str.isdigit(P) or P == '':
        return True
    else:
        return False

def get_all_results(selected_date_var):
    connection = connect_to_database()
    cursor = connection.cursor()

    selected_date = selected_date_var.get()
    date_range = selected_date.split(" ~ ")
    start_date = datetime.strptime(date_range[0], "%Y/%m/%d").date()
    end_date = datetime.strptime(date_range[1], "%Y/%m/%d").date()

    query = f"SELECT t.SN, t.SnapTime, t.PaintNum, d.PaintNum AS Paint1, t.Result " \
            f"FROM steel_paint_test AS t " \
            f"LEFT JOIN steel_paint_detail AS d ON t.SN = d.MasterSN AND d.PaintPos = 'Paint1' " \
            f"WHERE t.SnapTime BETWEEN '{start_date.strftime('%Y/%m/%d')}' AND '{end_date.strftime('%Y/%m/%d')}'"
    cursor.execute(query)
    result = cursor.fetchall()
    result = sorted(result, key=lambda x: x[1], reverse=True)

    cursor.close()
    connection.close()

    return result

def next_page(total_pages_label, tree, selected_date_var):
    current_page = int(total_pages_label.cget("text").split("/")[0])
    total_pages = int(total_pages_label.cget("text").split("/")[1])
    
    if current_page == total_pages:
        messagebox.showinfo("提示", "已經是最末頁")
        return
    elif current_page < total_pages:
        results = get_all_results(selected_date_var)
    
    current_page += 1
    total_pages_label.config(text=f"{current_page}/{total_pages}")

    tree.delete(*tree.get_children())

    for row in results[(current_page - 1) * 10:current_page * 10]:
        sn = row[0]
        snap_time = row[1]
        l1_data = row[2] if len(row) > 2 else ""
        cam1_num = row[3] if len(row) > 3 else ""
        result = row[4] if len(row) > 4 else ""

        if result == "OK":
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("OK",))
        elif result == "CHECK":
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("CHECK",))
        else:
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("NG",))

def previous_page(total_pages_label, tree, selected_date_var):
    current_page = int(total_pages_label.cget("text").split("/")[0])
    total_pages = int(total_pages_label.cget("text").split("/")[1])
    
    if current_page == 1:
        return
    elif current_page > 1:
        results = get_all_results(selected_date_var)
    
    current_page -= 1
    total_pages_label.config(text=f"{current_page}/{total_pages}")
    
    tree.delete(*tree.get_children())

    for row in results[(current_page - 1) * 10:current_page * 10]:
        sn = row[0]
        snap_time = row[1]
        l1_data = row[2] if len(row) > 2 else ""
        cam1_num = row[3] if len(row) > 3 else ""
        result = row[4] if len(row) > 4 else ""

        if result == "OK":
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("OK",))
        elif result == "CHECK":
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("CHECK",))
        else:
            tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("NG",))

def update_image_frames(sn, img_frame3, img_frame4, window_width, window_height):
    # 建立資料庫連接
    connection = connect_to_database()
    cursor = connection.cursor()
    
    # 查詢相關的資料
    query_cam1 = f"SELECT Image FROM steel_paint_image WHERE MasterSN = {sn} AND ImageType = 'Cam1'"
    cursor.execute(query_cam1)
    result_cam1 = cursor.fetchone()
    image_cam1 = result_cam1[0] if result_cam1 is not None else None

    query_paint1 = f"SELECT Image FROM steel_paint_detail WHERE MasterSN = {sn} AND PaintPos = 'Paint1'"
    cursor.execute(query_paint1)
    result_paint1 = cursor.fetchone()
    image_paint1 = result_paint1[0] if result_paint1 is not None else None

    # 關閉資料庫連接
    cursor.close()
    connection.close()
    
    # 固定框架的大小並顯示圖像

    img_frame3.config(width=int(window_width * 0.23), height=int(window_height * 0.32))
    show_image_in_frame(image_cam1, img_frame3)

    img_frame4.config(width=int(window_width * 0.23), height=int(window_height * 0.32))
    show_image_in_frame(image_paint1, img_frame4)

def update_var_labels(sn, var_label3, var_label4):
    # 建立資料庫連接
    connection = connect_to_database()
    cursor = connection.cursor()
    
    # 查詢相關的資料
    query_paint_num = f"SELECT PaintNum FROM steel_paint_test WHERE SN = {sn}"
    cursor.execute(query_paint_num)
    result_paint_num = cursor.fetchone()
    L1data = result_paint_num[0] if result_paint_num is not None else "NONE"
    var_label3.configure(text="")
    var_label3.configure(text=L1data)

    query_paint1_num = f"SELECT PaintNum FROM steel_paint_detail WHERE MasterSN = {sn} AND PaintPos = 'Paint1'"
    cursor.execute(query_paint1_num)
    result_paint1_num = cursor.fetchone()
    paint1 = result_paint1_num[0] if result_paint1_num is not None else "NONE"
    var_label4.configure(text="")
    var_label4.configure(text=paint1)
    '''
    query_paint2_num = f"SELECT PaintNum FROM steel_paint_detail WHERE MasterSN = {sn} AND PaintPos = 'Paint2'"
    cursor.execute(query_paint2_num)
    result_paint2_num = cursor.fetchone()
    paint2 = result_paint2_num[0] if result_paint2_num is not None else "NONE"
    var_label4.configure(text="")
    var_label4.configure(text=paint2)
    '''
    # 關閉資料庫連接
    cursor.close()
    connection.close()

def show_image_in_frame(image_data, frame):
    # 清除框架中的舊圖像
    for widget in frame.winfo_children():
        widget.destroy()

    if image_data == b'None':
        image = Image.open("chong.png")
    else:
        # 將base64編碼的圖像數據轉換回二進制數據
        image_data = base64.b64decode(image_data)

        # 將圖像數據轉換為PIL Image對象
        image = Image.open(BytesIO(image_data))

    # 根據框架的大小調整圖像大小
    img_width, img_height = 450, 350
    image = image.resize((img_width, img_height), Image.LANCZOS)

    # 將圖像轉換為PhotoImage對象
    photo = ImageTk.PhotoImage(image)

    # 在框架中創建圖像標籤並顯示圖像
    image_label = ttk.Label(frame, image=photo)
    image_label.image = photo  # 保留對圖像的引用，以避免垃圾回收
    image_label.pack()

def select_row(tree, event, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, window_width, window_height):
        # 獲取選取的行
        selected_item = tree.focus()
        # 獲取行的數據
        values = tree.item(selected_item)['values']
        if values and len(values) > 0:
            # 更新圖像框和標籤的內容
            update_image_frames(values[0], img_frame3, img_frame4, window_width, window_height)
            update_var_labels(values[0], var_label3, var_label4)

def export_data_to_excel(selected_date_var):
    # 獲取日期範圍內的資料
    data = get_all_results(selected_date_var)

    # 建立 DataFrame
    df = pd.DataFrame(data, columns=["編號", "日期", "L1資料", "Cam1號碼", "比對結果"])

    # 將日期欄位轉換為 datetime 格式
    df["日期"] = pd.to_datetime(df["日期"])

    # 使用者選擇要存的路徑和檔案名稱
    root = tk.Tk()
    root.withdraw()
    today = date.today().strftime("%Y%m%d")

    # 指定預設檔案名稱為當天日期
    default_filename = f"鋼捲噴字記錄_{today}.xlsx"
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=default_filename)
    root.destroy()

    # 建立新的 Excel 工作簿
    workbook = Workbook()
    sheet = workbook.active

    # 設定標題欄位寬度
    header_widths = [9, 20, 12, 12, 9]
    for col_num, width in enumerate(header_widths, 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = width

    # 寫入標題
    headers = ["編號", "日期", "L1資料", "Cam1號碼", "比對結果"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)

    # 寫入資料
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            if isinstance(cell_value, date):
                cell.value = cell_value
                cell.number_format = 'yyyy/mm/dd'  # 設定日期格式
            else:
                cell.value = cell_value

    # 儲存 Excel 檔案
    workbook.save(save_path)
    messagebox.showinfo("提示", "資料匯出完成。")

def show_enlarge_view(img_frame3, img_frame4, tree):
    global enlarge_view_window  # 使用全局變數
    if enlarge_view_window and enlarge_view_window.winfo_exists():
        # 如果已經存在放大畫面視窗，則將焦點轉移到該視窗
        enlarge_view_window.lift()
    else:
        # 建立放大畫面視窗
        enlarge_view_window = tk.Toplevel()
        enlarge_view_window.title("放大畫面")
        enlarge_view_window.geometry("1620x650")
        enlarge_view_window.resizable(False, False)
    def show_image_in_frame_(base64_image_data, label_widget):
        if base64_image_data:
            if base64_image_data == b'None':
                image = Image.open("chong.png")
            else:
                # 將base64編碼的圖像數據轉換回二進制數據
                image_data = base64.b64decode(base64_image_data)

                # 將圖像數據轉換為PIL Image對象
                image = Image.open(BytesIO(image_data))

            img_width, img_height = 800, 600
            image = image.resize((img_width, img_height), Image.LANCZOS)
            # 將PIL Image轉換為PhotoImage對象
            photo_image = ImageTk.PhotoImage(image)

            # 在標籤控件中顯示圖像
            label_widget.configure(image=photo_image)
            label_widget.image = photo_image
        else:
            # 如果沒有圖像數據，則將標籤控件的圖像設置為空
            label_widget.configure(image=None)
    # 建立第一個frame，用於顯示img_frame3的放大畫面
    frame1 = ttk.Frame(enlarge_view_window, width=800, height=600)
    frame1.pack(side=tk.LEFT, padx=10, pady=10)

    # 建立第二個frame，用於顯示img_frame4的放大畫面
    frame2 = ttk.Frame(enlarge_view_window, width=800, height=600)
    frame2.pack(side=tk.LEFT, padx=10, pady=10)

    image_label3 = ttk.Label(frame1)
    image_label3.pack()
    image_label4 = ttk.Label(frame2)
    image_label4.pack()

    # 獲取選取的行
    selected_item = tree.focus()
    
    # 只有在選擇了行時才執行顯示圖片的功能
    if selected_item:
        # 建立第一個frame，用於顯示img_frame3的放大畫面
        frame1 = ttk.Frame(enlarge_view_window, width=800, height=600)
        frame1.pack(side=tk.LEFT, padx=10, pady=10)

        # 建立第二個frame，用於顯示img_frame4的放大畫面
        frame2 = ttk.Frame(enlarge_view_window, width=800, height=600)
        frame2.pack(side=tk.LEFT, padx=10, pady=10)

        image_label3 = ttk.Label(frame1)
        image_label3.pack()
        image_label4 = ttk.Label(frame2)
        image_label4.pack()

        # 獲取行的數據
        values = tree.item(selected_item)['values']
        # 建立資料庫連接
        connection = connect_to_database()
        cursor = connection.cursor()

        # 查詢相關的資料
        query_cam1 = f"SELECT Image FROM steel_paint_image WHERE MasterSN = {values[0]} AND ImageType = 'Cam1'"
        cursor.execute(query_cam1)
        result_cam1 = cursor.fetchone()
        image_cam1 = result_cam1[0] if result_cam1 is not None else None

        query_paint1 = f"SELECT Image FROM steel_paint_detail WHERE MasterSN = {values[0]} AND PaintPos = 'Paint1'"
        cursor.execute(query_paint1)
        result_paint1 = cursor.fetchone()
        image_paint1 = result_paint1[0] if result_paint1 is not None else None

        # 關閉資料庫連接
        cursor.close()
        connection.close()

        # 將圖片放大到所需尺寸並顯示在對應的框架中
        if image_cam1:
            show_image_in_frame_(image_cam1, image_label3)  # 直接傳遞base64編碼的圖片數據
        else:
            show_image_in_frame_(None, image_label3)  # 如果圖片為空，則清空label_widget中的圖片

        if image_paint1:
            show_image_in_frame_(image_paint1, image_label4)  # 直接傳遞base64編碼的圖片數據
        else:
            show_image_in_frame_(None, image_label4)  # 如果圖片為空，則清空label_widget中的圖片

def flash_window(window):
    current_bg = window.cget('background')
    next_bg = 'white' if current_bg == 'red' else 'red'
    window.configure(background=next_bg)
    window.after(1000, flash_window, window)

def update_database_result(selected_sn, result):
    # 連接資料庫
    connection = connect_to_database()
    cursor = connection.cursor()

    # 更新比對結果
    sql = f"UPDATE steel_paint_test SET Result = '{result}' WHERE SN = '{selected_sn}'"

    try:
        cursor.execute(sql)
        connection.commit()
        logging.info(f"比對結果已更新到資料庫，編號'{selected_sn}'比對結果改為'{result}'")
    except Exception as e:
        connection.rollback()
        #print("更新比對結果時發生錯誤:", str(e))
        logging.error(f'更新比對結果時發生錯誤: {e}')
    finally:
        cursor.close()
        connection.close()
        
def update_result(root,selected_date_var,query_entry,total_pages_label,tree, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2):
    selected_item = tree.focus()  # 獲取選取的行
    values = tree.item(selected_item)['values']  # 獲取行的數據

    if values and len(values) > 0:
        selected_sn = values[0]  # 選取行的 SN
        selected_result = ""  # 選取行的比對結果

        if selected_sn != "":
            # 建立新視窗
            update_window = tk.Toplevel(root)
            update_window.title("更改比對結果")
            update_window.geometry("600x200")
            update_window.resizable(False, False)

            # 設定視窗內容
            label = tk.Label(update_window, text="選擇結果類型", font=("Arial", 20))
            label.pack()

            # 按鈕事件處理函數
            def update_selected_result(result):
                # 更新資料庫中的比對結果
                update_database_result(selected_sn, result)
                # 關閉視窗
                update_window.destroy()
                # 更新當前頁面的結果
                renew(selected_date_var,query_entry,total_pages_label,tree, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2)

            # 加入按鈕
            button_frame = tk.Frame(update_window)
            button_frame.pack()

            style = ttk.Style()
            style.configure("TButton", font=("Arial", 20))

            ok_button = ttk.Button(button_frame, text="OK", style="TButton",
                                   command=lambda: update_selected_result("OK"))
            ok_button.pack(side=tk.LEFT, padx=5, pady=5)

            check_button = ttk.Button(button_frame, text="CHECK", style="TButton",
                                      command=lambda: update_selected_result("CHECK"))
            check_button.pack(side=tk.LEFT, padx=5, pady=5)

            ng_button = ttk.Button(button_frame, text="NG", style="TButton",
                                   command=lambda: update_selected_result("NG"))
            ng_button.pack(side=tk.LEFT, padx=5, pady=5)

def update(selected_date_var,tree,total_pages_label,root,query_entry, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2):
    global original_result_count
    global notification  
    global ng_Checked

    if original_result_count == 0:
        original_result_count = len(get_all_results(selected_date_var))
        print("original_result_count", original_result_count)

    # 持續追蹤資料庫內容
    results = get_all_results(selected_date_var)

    # 檢查是否有新資料
    if len(results) > original_result_count:
        ng_Checked = False
        # 更新原始結果數量
        original_result_count = 0
        # 清空表格內容
        tree.delete(*tree.get_children())
        # 更新頁數標籤
        total_pages = (len(results) - 1) // 10 + 1
        current_page = int(total_pages_label.cget("text").split("/")[0])
        total_pages_label.config(text=f"{current_page}/{total_pages}")
        #更新表格內容
        for row in results[(current_page - 1) * 10:current_page * 10]:
            sn = row[0]
            snap_time = row[1]
            l1_data = row[2] if len(row) > 2 else ""
            cam1_num = row[3] if len(row) > 3 else ""
            result = row[4] if len(row) > 4 else ""

            if result == "OK":
                tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("OK",))
            elif result == "CHECK":
                tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("CHECK",))
            else:
                tree.insert("", tk.END, values=(sn, snap_time, l1_data, cam1_num, result), tags=("NG",))
        '''
        # 檢查第一筆資料的比對結果
        if results[0][4] == "NG" and notification == False and ng_Checked == False:
            notification = True
            # 更新頁數標籤
            total_pages = (len(results) - 1) // 10 + 1
            current_page = 1
            total_pages_label.config(text=f"{current_page}/{total_pages}")
            switch_page(current_page, total_pages_label, tree, selected_date_var)
            # 選取第一行資料
            tree.selection_set(tree.get_children()[0])  # 選取第一筆資料
            tree.focus(tree.get_children()[0])  # 設定焦點於第一筆資料
            tree.see(tree.get_children()[0])  # 滾動表格以確保選取的資料可見
            # 建立新視窗
            update_window = tk.Toplevel(root)
            update_window.title("辨識結果不合")
            update_window.geometry("910x400")
            update_window.resizable(False, False)

            # 設定視窗內容
            label = tk.Label(update_window, text="辨識結果不合。\n例外狀況請按確認\n錯誤請按錯誤", font=("Arial", 20))
            label.pack()

            # 加入按鈕
            button_frame = tk.Frame(update_window)
            button_frame.pack()

            style = ttk.Style()
            style.configure("TButton", font=("Arial", 20))

            def confirm_action():
                global notification
                selected_item = tree.focus()
                selected_sn = tree.item(selected_item)['values'][0]
                update_database_result(selected_sn, "CHECK")
                notification = False
                update_window.destroy()
                renew(selected_date_var,query_entry,total_pages_label,tree, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2)

            def ng_action():
                global notification
                global ng_Checked
                notification = False
                ng_Checked = True
                update_window.destroy()
                renew(selected_date_var,query_entry,total_pages_label,tree, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2)

            check_button = ttk.Button(button_frame, text="確認", style="TButton", command=confirm_action)
            check_button.pack(side=tk.LEFT, padx=5, pady=5)

            ng_button = ttk.Button(button_frame, text="錯誤", style="TButton", command=ng_action)
            ng_button.pack(side=tk.LEFT, padx=5, pady=5)

            flash_window(update_window) 
            '''
    root.after(1000, lambda: update(selected_date_var,tree,total_pages_label,root,query_entry, img_frame1, img_frame2, img_frame3, img_frame4, var_label2, var_label3, var_label4, image_label2))


def display_image_on_canvas(canvas, image_path):
    def resize_image(image_path, width, height):
        # 使用PIL庫來載入圖片並進行resize
        image = Image.open(image_path)
        resized_image = image.resize((width, height), Image.LANCZOS)
        return ImageTk.PhotoImage(resized_image)
    # 調整圖片大小
    resized_image = resize_image(image_path, 450, 350)
    # 在Canvas上建立圖片物件
    canvas.create_image(0, 0, anchor=tk.NW, image=resized_image)
    # 將圖片物件儲存到Canvas物件中
    canvas.image = resized_image
